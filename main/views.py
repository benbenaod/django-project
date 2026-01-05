from __future__ import annotations

from pathlib import Path
import json
import os
from typing import Dict, Iterable, List, Optional, Tuple

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, get_user_model, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils.html import escape
from django.views.decorators.http import require_GET, require_POST

from openpyxl import load_workbook

from .forms import CourseForm
from .models import Course, Student, Teacher


# ==============================
# Google Map building links
# ==============================
BUILDING_URL_MAP = {
    "F": "https://www.google.com/maps/place/%E5%9C%8B%E7%AB%8B%E8%87%BA%E5%8C%97%E8%AD%B7%E7%90%86%E5%81%A5%E5%BA%B7%E5%A4%A7%E5%AD%B8%E5%AD%B8%E6%80%9D%E6%A8%93/@25.1186186,121.5166288,17z/data=!3m1!4b1!4m6!3m5!1s0x3442af4ac9da7987:0xf36d626d63834f5!8m2!3d25.1186138!4d121.5192037!16s%2Fg%2F11s82z2lrp?entry=ttu&g_ep=EgoyMDI1MTIwOS4wIKXMDSoKLDEwMDc5MjA2OUgBUAM%3D",
    "S": "https://www.google.com/maps/place/%E5%9C%8B%E7%AB%8B%E8%87%BA%E5%8C%97%E8%AD%B7%E7%90%86%E5%81%A5%E5%BA%B7%E5%A4%A7%E5%AD%B8%E7%A7%91%E6%8A%80%E5%A4%A7%E6%A8%93/@25.117542,121.5180909,17z/data=!3m1!5s0x3442ae8a4f198def:0x16fcf46afefac4c2!4m16!1m9!3m8!1s0x3442ae8967e29825:0xa74a929b7ae3dbf6!2z5ZyL56uL6Ie65YyX6K2355CG5YGl5bq35aSn5a2456eR5oqA5aSn5qiT!8m2!3d25.1175372!4d121.5206658!9m1!1b1!16s%2Fg%2F11b6jgqh03!3m5!1s0x3442ae8967e29825:0xa74a929b7ae3dbf6!8m2!3d25.1175372!4d121.5206658!16s%2Fg%2F11b6jgqh03?entry=ttu&g_ep=EgoyMDI1MTIwOS4wIKXMDSoKLDEwMDc5MjA2OUgBUAM%3D",
    "B": "https://www.google.com/maps/place/%E5%9C%8B%E7%AB%8B%E8%87%BA%E5%8C%97%E8%AD%B7%E7%90%86%E5%81%A5%E5%BA%B7%E5%A4%A7%E5%AD%B8%E8%A6%AA%E4%BB%81%E6%A8%93/@25.1185795,121.5185797,17z/data=!3m2!4b1!5s0x3442ae8a4f198def:0x16fcf46afefac4c2!4m6!3m5!1s0x3442af851c386faf:0xc3edb631a5715fd3!8m2!3d25.1185747!4d121.5211546!16s%2Fg%2F11ryljg7x2?entry=ttu&g_ep=EgoyMDI1MTIwOS4wIKXMDSoKLDEwMDc5MjA2OUgBUAM%3D",
    "G": "https://www.google.com.tw/maps/place/%E5%9C%8B%E7%AB%8B%E8%87%BA%E5%8C%97%E8%AD%B7%E7%90%86%E5%81%A5%E5%BA%B7%E5%A4%A7%E5%AD%B8%E6%A0%A1%E6%9C%AC%E9%83%A8/@25.1175841,121.5166108,17z/data=!3m2!4b1!5s0x3442ae8a4f198def:0x16fcf46afefac4c2!4m6!3m5!1s0x3442ae8bc54ebc79:0xfd2a9d659e97b078!8m2!3d25.1175793!4d121.5214817!16s%2Fm%2F0z8mtpb?entry=ttu&g_ep=EgoyMDI1MTIwOS4wIKXMDSoASAFQAw%3D%3D",
}


# ==============================
# 常用安全工具
# ==============================
def safe_str(v) -> str:
    """任何值 → 安全字串（處理 None/NaN/'nan'）"""
    if v is None:
        return ""
    try:
        import math

        if isinstance(v, float) and math.isnan(v):
            return ""
    except Exception:
        pass
    s = str(v).strip()
    return "" if s.lower() == "nan" else s


def esc(v) -> str:
    return escape(safe_str(v))


def safe_get(row: dict, col_name: str, default="") -> str:
    try:
        return safe_str(row.get(col_name, default))
    except Exception:
        try:
            return safe_str(row[col_name])
        except Exception:
            return safe_str(default)


# ==============================
# Excel 資料夾與讀取
# ==============================
HEADER_ROW = 5  # 你的 pandas header=4 => Excel 第 5 列是欄名
BATCH_SIZE = 300  # Render 小方案建議 200~500


def get_excel_dir() -> Path:
    base = Path(settings.BASE_DIR)
    p = base  # 你目前把 xlsx 放 repo 根目錄
    xlsx_files = list(p.glob("*.xlsx"))
    if xlsx_files:
        print(f"✅ 使用 Excel 資料夾：{p}")
        print("✅ 找到 xlsx：", [f.name for f in xlsx_files])
        return p
    print(f"⚠️ 在 {p} 裡沒有找到任何 .xlsx 檔案")
    return p


EXCEL_DIR = get_excel_dir()


def _iter_xlsx_dict_rows(
    file_path: Path, header_row: int = HEADER_ROW
) -> Tuple[Optional[List[str]], Iterable[dict]]:
    """
    逐列讀取 xlsx，回傳 (headers, rows_generator)。
    重點：generator 結束後會自動 wb.close()，避免檔案 handle/記憶體累積。
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)

    try:
        # 跳到 header_row
        for _ in range(header_row - 1):
            next(it, None)

        headers = next(it, None)
        if not headers:
            wb.close()
            return None, []

        headers = [safe_str(h) for h in headers]
        col_idx = {h: i for i, h in enumerate(headers) if h}

        def gen():
            try:
                for values in it:
                    row = {}
                    for name, i in col_idx.items():
                        row[name] = values[i] if i < len(values) else None
                    yield row
            finally:
                wb.close()

        return headers, gen()

    except Exception:
        wb.close()
        raise


# ==============================
# 教室欄位統一
# ==============================
ROOM_COL_CANDIDATES = [
    "上課地點",
    "教室地點",
    "上課教室",
    "教室",
    "上課位置",
    "上課地點(教室)",
    "上課教室地點",
    "上課地點/教室",
    "地點",
    "位置",
]


def room_from_row(row) -> str:
    for col in ROOM_COL_CANDIDATES:
        v = safe_get(row, col)
        if v:
            return v
    return ""


def room_display(c: Course) -> str:
    v = safe_str(getattr(c, "classroom", ""))
    v = v or safe_str(getattr(c, "room", ""))
    v = v or safe_str(getattr(c, "location", ""))
    return v or "-"


def building_url_from_room(room: str) -> str:
    """
    依教室字串推測大樓：如 F308 / S210 / B101 / G... 的第一碼
    回傳對應 BUILDING_URL_MAP，找不到回傳空字串
    """
    r = safe_str(room).strip()
    if not r or r == "-":
        return ""
    key = r.strip().upper()[:1]
    return safe_str(BUILDING_URL_MAP.get(key, ""))


# ==============================
# Teacher meta：中文姓名 / 類別 / 分機
# ==============================
def _teacher_meta_from_obj(t: Optional[Teacher]) -> Tuple[str, str, str]:
    if not t:
        return "", "", ""

    name_ch = safe_str(getattr(t, "name_ch", ""))

    category = (
        safe_str(getattr(t, "category", ""))
        or safe_str(getattr(t, "type", ""))
        or safe_str(getattr(t, "role", ""))
        or safe_str(getattr(t, "title", ""))
    )

    ext = (
        safe_str(getattr(t, "extension", ""))
        or safe_str(getattr(t, "ext", ""))
        or safe_str(getattr(t, "phone_ext", ""))
        or safe_str(getattr(t, "school_ext", ""))
        or safe_str(getattr(t, "office_ext", ""))
    )

    return name_ch, category, ext


def teacher_meta_for_course(c: Optional[Course]) -> Tuple[str, str, str]:
    if not c:
        return "", "", ""

    t = getattr(c, "teacher_ref", None)

    if not t:
        tname = safe_str(getattr(c, "teacher", ""))
        if tname:
            t = Teacher.objects.filter(name_ch=tname).first()

    name_ch, category, ext = _teacher_meta_from_obj(t)

    if not name_ch:
        name_ch = safe_str(getattr(c, "teacher", ""))

    return name_ch, category, ext


# ==============================
# 系所代碼 → 中文系所名
# ==============================
DEPT_NAME_MAP = {
    "22140": "資訊管理系",
    "22160": "資訊管理系碩士班",
    "11140": "護理系",
    "21140": "健康事業管理系",
    "24120": "二年制長期照護系",
    "13140": "高齡健康照護系",
    "31140": "嬰幼兒保育系",
    "25140": "語言治療與聽力學系",
    "23140": "休閒產業與健康促進系",
    "32140": "運動保健系",
    "33140": "生死與健康心理諮商系",
    # ...（你原本那一大段照貼即可）
}

FOUR_TECH_DEPTS = {
    "22140": "資訊管理系",
    "22160": "資訊管理系碩士班",
    "11140": "護理系",
    "21140": "健康事業管理系",
    "24120": "二年制長期照護系",
    "13140": "高齡健康照護系",
    "31140": "嬰幼兒保育系",
    "25140": "語言治療與聽力學系",
    "23140": "休閒產業與健康促進系",
    "32140": "運動保健系",
    "33140": "生死與健康心理諮商系",
}


def dept_display(code: str) -> str:
    code = safe_str(code)
    return DEPT_NAME_MAP.get(code, "") or code or "-"


def apply_system_filter(qs, system_value: str):
    system_value = safe_str(system_value)
    if not system_value:
        return qs

    def hit(keyword: str) -> Q:
        keyword = safe_str(keyword)
        if not keyword:
            return Q()
        return (
            Q(system__icontains=keyword)
            | Q(schedule_old_name__icontains=keyword)
            | Q(schedule_old_code__icontains=keyword)
            | Q(class_group__icontains=keyword)
            | Q(teaching_group__icontains=keyword)
            | Q(course_name__icontains=keyword)
            | Q(department_code__icontains=keyword)
        )

    def dept_codes_by_keywords(keywords):
        codes = set()
        for code, name in DEPT_NAME_MAP.items():
            for k in keywords:
                if k and k in (name or ""):
                    codes.add(code)
                    break
        return codes

    if system_value == "二專":
        keywords = ["1D110"]
        codes = dept_codes_by_keywords(keywords)
        q = hit("1D110")
        if codes:
            q |= Q(department_code__in=list(codes))
        return qs.filter(q)

    if system_value == "二技":
        keywords = ["二年制"]
        codes = dept_codes_by_keywords(keywords)
        q = hit("二年制") | hit("二技")
        if codes:
            q |= Q(department_code__in=list(codes))
        return qs.filter(q)

    if system_value == "二技(三年)":
        keywords = ["二年制進修部"]
        codes = dept_codes_by_keywords(keywords)
        q = hit("二年制進修部") | hit("二技(三年)")
        if codes:
            q |= Q(department_code__in=list(codes))
        return qs.filter(q)

    if system_value == "四技":
        keywords = ["四年制", "四技"]
        codes = dept_codes_by_keywords(keywords) | set(FOUR_TECH_DEPTS.keys())
        q = hit("四年制") | hit("四技") | Q(department_code__in=list(codes))
        return qs.filter(q)

    if system_value == "學士後多元專長":
        keywords = ["學士後多元專長"]
        codes = dept_codes_by_keywords(keywords)
        q = hit("學士後多元專長")
        if codes:
            q |= Q(department_code__in=list(codes))
        return qs.filter(q)

    if system_value == "碩士班":
        keywords = ["研究所", "學生專班", "碩士班", "碩士在職"]
        codes = dept_codes_by_keywords(keywords)
        q = Q()
        for k in keywords:
            q |= hit(k)
        if codes:
            q |= Q(department_code__in=list(codes))
        return qs.filter(q)

    if system_value == "博士班":
        keywords = ["博士"]
        codes = dept_codes_by_keywords(keywords)
        q = hit("博士")
        if codes:
            q |= Q(department_code__in=list(codes))
        return qs.filter(q)

    if system_value == "學士後學位學程":
        keywords = ["學士後教保學位學程"]
        codes = dept_codes_by_keywords(keywords)
        q = hit("學士後教保學位學程")
        if codes:
            q |= Q(department_code__in=list(codes))
        return qs.filter(q)

    if system_value == "學士後系":
        keywords = ["學士後學士班"]
        codes = dept_codes_by_keywords(keywords)
        q = hit("學士後學士班")
        if codes:
            q |= Q(department_code__in=list(codes))
        return qs.filter(q)

    return qs


# ==============================
# 權限/身分 helpers（✅ 一鍵登入/正常登入都能用的版本）
# ==============================
def get_user_display_name(user) -> Tuple[str, str]:
    if not user or not getattr(user, "is_authenticated", False):
        return "", ""

    # ✅ 超關鍵：staff/superuser 一律視為老師，避免 Teacher 沒綁 user 導致「登入了但不能用」
    if getattr(user, "is_superuser", False) or getattr(user, "is_staff", False):
        name = safe_str(getattr(user, "first_name", "")) or safe_str(getattr(user, "username", ""))
        return "老師", name

    t = Teacher.objects.filter(user=user).first()
    if t:
        name = (
            safe_str(getattr(t, "name_ch", ""))
            or safe_str(getattr(user, "first_name", ""))
            or safe_str(getattr(user, "username", ""))
        )
        return "老師", name

    s = Student.objects.filter(user=user).first()
    if s:
        name = (
            safe_str(getattr(s, "name", ""))
            or safe_str(getattr(user, "first_name", ""))
            or safe_str(getattr(user, "username", ""))
        )
        return "學生", name

    first = safe_str(getattr(user, "first_name", ""))
    if first:
        return "使用者", first

    return "使用者", safe_str(getattr(user, "username", ""))


def is_teacher_admin(request) -> bool:
    if not request.user.is_authenticated:
        return False
    # ✅ staff/superuser 直接當老師
    if request.user.is_staff or request.user.is_superuser:
        return True
    return Teacher.objects.filter(user=request.user).exists()


def is_student_user(request) -> bool:
    return bool(request.user.is_authenticated and Student.objects.filter(user=request.user).exists())


def ensure_role_profile(user, role: str):
    """
    ✅ 關鍵：登入成功後，如果 Teacher/Student 沒綁 user，就自動補齊
    role: "student" / "admin"
    """
    role = safe_str(role)

    if role == "student":
        if Student.objects.filter(user=user).exists():
            return

        sid = safe_str(getattr(user, "username", "")) or "demo"
        try:
            obj = Student.objects.create(user=user, student_id=sid)
        except Exception:
            # 如果你的 Student model 沒有 student_id 或有其他限制，就退到最小建立
            obj = Student.objects.create(user=user)

        if hasattr(obj, "name") and not safe_str(getattr(obj, "name", "")):
            obj.name = safe_str(getattr(user, "first_name", "")) or safe_str(getattr(user, "username", ""))
            try:
                obj.save(update_fields=["name"])
            except Exception:
                obj.save()
        return

    if role == "admin":
        if not (user.is_staff or user.is_superuser):
            user.is_staff = True
            user.save(update_fields=["is_staff"])

        if Teacher.objects.filter(user=user).exists():
            return

        tname = safe_str(getattr(user, "first_name", "")) or safe_str(getattr(user, "username", "")) or "老師"
        try:
            Teacher.objects.create(user=user, name_ch=tname)
        except Exception:
            Teacher.objects.create(user=user)
        return


# ==============================
# DEMO / 預設帳密（可用環境變數在 Render 開啟）
# ==============================
_DEFAULT_CREATED = False

DEMO_AUTO_LOGIN = os.environ.get("DEMO_AUTO_LOGIN", "0") == "1"
DEMO_SEED_ACCOUNTS = os.environ.get("DEMO_SEED_ACCOUNTS", "0") == "1"
DEFAULT_STUDENT_USERNAME = "ben"
DEFAULT_TEACHER_USERNAME = "dora"


def demo_auto_login(request):
    """
    DEMO 用：不用帳密直接登入。
    - 需設定環境變數 DEMO_AUTO_LOGIN=1 才會啟用
    - 預設自動登入學生 ben
    - ?as=teacher 可切換老師 dora
    - ?as=student 切回學生 ben
    """
    if not DEMO_AUTO_LOGIN:
        return
    if request.user.is_authenticated:
        return

    as_role = safe_str(request.GET.get("as"))  # student / teacher
    username = DEFAULT_STUDENT_USERNAME
    role_for_profile = "student"

    if as_role == "teacher":
        username = DEFAULT_TEACHER_USERNAME
        role_for_profile = "admin"
    elif as_role == "student":
        username = DEFAULT_STUDENT_USERNAME
        role_for_profile = "student"

    User = get_user_model()
    user = User.objects.filter(username=username).first()
    if not user:
        return

    # ✅ 補齊身分資料，避免「登入了但功能不能用」
    ensure_role_profile(user, role_for_profile)

    # ✅ 免密碼登入要指定 backend
    login(request, user, backend="django.contrib.auth.backends.ModelBackend")


def ensure_default_accounts():
    """
    ✅ 關鍵修正：
    1) 不只 DEBUG，Render 也能用（用環境變數 DEMO_SEED_ACCOUNTS=1 或 DEMO_AUTO_LOGIN=1 開啟）
    2) 就算 user 已存在，也會「確保密碼是正確的」(避免你以前測試留下舊密碼)
    """
    global _DEFAULT_CREATED
    if _DEFAULT_CREATED:
        return

    allow = bool(getattr(settings, "DEBUG", False)) or DEMO_SEED_ACCOUNTS or DEMO_AUTO_LOGIN
    if not allow:
        return

    User = get_user_model()

    DEFAULT_ACCOUNTS = [
        {"role": "teacher", "username": "dora", "password": "a", "teacher_name": "中岳"},
        {"role": "student", "username": "ben", "password": "a", "student_id": "122214132", "student_name": "童國原"},
    ]

    for item in DEFAULT_ACCOUNTS:
        username = safe_str(item.get("username"))
        password = item.get("password") or ""
        if not username:
            continue

        user, _created = User.objects.get_or_create(username=username)

        # ✅ 核心：就算不是 created，也強制確保密碼對
        if not user.check_password(password):
            user.set_password(password)

        # 老師帳號：給 is_staff（你用 admin 身分判斷會更穩）
        if item["role"] == "teacher":
            if not user.is_staff:
                user.is_staff = True

        # 塞 first_name 方便顯示
        if not safe_str(getattr(user, "first_name", "")):
            display = safe_str(item.get("teacher_name") or item.get("student_name") or username)
            user.first_name = display

        user.save()

        if item["role"] == "teacher":
            teacher_name = safe_str(item.get("teacher_name")) or username

            # 先綁「同名但 user 空」的 Teacher
            t = Teacher.objects.filter(name_ch=teacher_name, user__isnull=True).first()
            if t:
                t.user = user
                t.save(update_fields=["user"])
            else:
                Teacher.objects.get_or_create(user=user, defaults={"name_ch": teacher_name})

            Teacher.objects.filter(user=user).update(name_ch=teacher_name)

        elif item["role"] == "student":
            sid = safe_str(item.get("student_id")) or username
            sname = safe_str(item.get("student_name")) or username

            s = Student.objects.filter(student_id=sid).first()
            if s:
                if getattr(s, "user_id", None) is None:
                    s.user = user
                    s.save(update_fields=["user"])
            else:
                obj, _ = Student.objects.get_or_create(user=user, defaults={"student_id": sid})
                if hasattr(obj, "student_id") and not safe_str(getattr(obj, "student_id", "")):
                    obj.student_id = sid
                if hasattr(obj, "name") and not safe_str(getattr(obj, "name", "")):
                    obj.name = sname
                obj.save()

    _DEFAULT_CREATED = True


# ==============================
# Excel → Course 匯入
# ==============================
def _import_xlsx_to_course(file_path: Path) -> int:
    headers, rows = _iter_xlsx_dict_rows(file_path, header_row=HEADER_ROW)
    if not headers or rows is None:
        print(f"⚠️ 無法讀取表頭：{file_path.name}")
        return 0

    if "科目中文名稱" not in headers:
        print("⚠️ Excel 裡找不到『科目中文名稱』欄位，請確認欄位名稱。")
        print("目前欄位：", headers)
        return 0

    count = 0
    teacher_cache: Dict[str, Teacher] = {}
    batch: List[Course] = []

    def flush():
        nonlocal batch
        if not batch:
            return
        Course.objects.bulk_create(batch, batch_size=BATCH_SIZE)
        batch = []

    with transaction.atomic():
        for row in rows:
            course_name = safe_get(row, "科目中文名稱")
            if not course_name:
                continue

            teacher_name = safe_get(row, "主開課教師姓名")
            teacher_obj = None

            if teacher_name:
                teacher_obj = teacher_cache.get(teacher_name)
                if teacher_obj is None:
                    teacher_obj, _ = Teacher.objects.get_or_create(
                        name_ch=teacher_name,
                        defaults={"name_en": ""},
                    )
                    teacher_cache[teacher_name] = teacher_obj

            classroom_val = room_from_row(row)

            batch.append(
                Course(
                    number=safe_get(row, "編號"),
                    semester=safe_get(row, "學期"),
                    teacher=teacher_name,
                    course_code=safe_get(row, "科目代碼(新碼全碼)"),
                    department_code=safe_get(row, "系所代碼"),
                    core_code=safe_get(row, "核心四碼"),
                    group_code=safe_get(row, "科目組別"),
                    grade=safe_get(row, "年級"),
                    class_group=safe_get(row, "上課班組"),
                    course_name=course_name,
                    division=safe_get(row, "課別名稱"),
                    system=safe_get(row, "學制別"),
                    teaching_group=safe_get(row, "授課群組"),
                    week_info=safe_get(row, "上課週次"),
                    day=safe_get(row, "上課星期"),
                    period=safe_get(row, "上課節次"),
                    classroom=classroom_val,
                    course_summary_ch=safe_get(row, "課程中文摘要"),
                    course_summary_en=safe_get(row, "課程英文摘要"),
                    teacher_old_code=safe_get(row, "主開課教師代碼(舊碼)"),
                    course_old_code=safe_get(row, "科目代碼(舊碼)"),
                    schedule_old_code=safe_get(row, "課表代碼(舊碼)"),
                    schedule_old_name=safe_get(row, "課表名稱(舊碼)"),
                    teacher_old_code2=safe_get(row, "授課教師代碼(舊碼)"),
                    teacher_ref=teacher_obj,
                )
            )
            count += 1

            if len(batch) >= BATCH_SIZE:
                flush()

        flush()

    return count


def ensure_courses_loaded():
    """
    強烈建議：不要每次 request 都匯入。
    你目前用環境變數 AUTO_IMPORT=1 才會匯入，這樣 OK。
    """
    if os.environ.get("AUTO_IMPORT", "0") != "1":
        return
    if Course.objects.exists():
        return

    excel_files = sorted(EXCEL_DIR.glob("*.xlsx"))
    if not excel_files:
        print(f"⚠️ 在 {EXCEL_DIR} 裡沒有找到任何 .xlsx 檔案")
        return

    print(f"🔄 資料表為空，開始匯入 Excel（共 {len(excel_files)} 個檔案）...")
    total = 0
    for file_path in excel_files:
        try:
            print(f"➡ 匯入 {file_path.name}")
            n = _import_xlsx_to_course(file_path)
            total += n
            print(f"✅ {file_path.name} 匯入 {n} 筆")
        except Exception as e:
            print(f"❌ 匯入 {file_path.name} 失敗：{e}")
    print(f"🎉 匯入完成，共 {total} 筆")


# ==============================
# 個人課表（Session）
# ==============================
SESSION_KEY_PERSONAL = "personal_course_ids"

DEFAULT_PERSONAL_SEMESTER = "1141"
DEFAULT_PERSONAL_CLASS_GROUP = "A0"
REQUIRED_DEPT_FOR_RESEARCH = "22140"
REQUIRED_KEYWORDS = ["系統分析", "研究概論"]


def _get_personal_ids(request) -> List[int]:
    ids = request.session.get(SESSION_KEY_PERSONAL, [])
    if not isinstance(ids, list):
        ids = []
    out: List[int] = []
    for x in ids:
        try:
            xi = int(x)
            if xi not in out:
                out.append(xi)
        except Exception:
            continue
    return out


def _set_personal_ids(request, ids: List[int]):
    request.session[SESSION_KEY_PERSONAL] = list(ids)
    request.session.modified = True


def get_required_personal_courses():
    base = Course.objects.filter(
        semester=DEFAULT_PERSONAL_SEMESTER,
        class_group__icontains=DEFAULT_PERSONAL_CLASS_GROUP,
    )
    rule = {
        "系統分析": {},
        "研究概論": {"department_code": REQUIRED_DEPT_FOR_RESEARCH},
    }
    return base, rule


def resolve_required_course_ids() -> Dict[str, int]:
    base, rule = get_required_personal_courses()
    required_ids: Dict[str, int] = {}
    for kw, extra in rule.items():
        qs = base.filter(course_name__icontains=kw)
        if extra.get("department_code"):
            qs = qs.filter(department_code__exact=extra["department_code"])
        c = qs.order_by("day", "period", "course_name").first()
        if c:
            required_ids[kw] = c.id
    return required_ids


def ensure_fixed_personal_courses(request):
    if not is_student_user(request):
        return

    required_map = resolve_required_course_ids()
    existing = _get_personal_ids(request)
    existing_set = set(existing)

    for kw in REQUIRED_KEYWORDS:
        rid = required_map.get(kw)
        if rid and rid not in existing_set:
            existing.append(rid)
            existing_set.add(rid)

    _set_personal_ids(request, existing)


def is_required_course_id(course_id: int) -> bool:
    req = resolve_required_course_ids()
    return course_id in set(req.values())


def required_remove_message(course_id: int) -> str:
    req = resolve_required_course_ids()
    inv = {v: k for k, v in req.items()}
    name = inv.get(course_id, "此課程")
    return f"{name} 為必修安排，無法移除。"


def parse_periods(period_raw: str) -> List[int]:
    raw = safe_str(period_raw)
    if not raw:
        return []
    raw = raw.replace("、", ",").replace(" ", "")
    out: List[int] = []
    for part in raw.split(","):
        if not part:
            continue
        if "-" in part:
            try:
                a, b = part.split("-")
                a = int(a)
                b = int(b)
                for p in range(min(a, b), max(a, b) + 1):
                    if p not in out:
                        out.append(p)
            except Exception:
                continue
        else:
            try:
                p = int(part)
                if p not in out:
                    out.append(p)
            except Exception:
                continue
    return out


def _course_slots(course: Course) -> set:
    d = safe_str(getattr(course, "day", ""))
    if not d:
        return set()
    ps = parse_periods(safe_str(getattr(course, "period", "")))
    return {f"{d}-{p}" for p in ps}


def _conflict_slots(existing_courses: List[Course], new_course: Course) -> List[str]:
    exist_slots = set()
    for c in existing_courses:
        exist_slots |= _course_slots(c)
    new_slots = _course_slots(new_course)
    return sorted(list(exist_slots & new_slots))


def _format_conflicts(conflicts: List[str]) -> str:
    day_map = {"1": "一", "2": "二", "3": "三", "4": "四", "5": "五", "6": "六", "7": "日"}
    items = []
    for k in conflicts:
        try:
            d, p = k.split("-")
        except Exception:
            continue
        items.append(f"星期{day_map.get(d, d)} 第{p}節")
    return "、".join(items)


# ==============================
# 課表 HTML（含個人課表）
# ==============================
def build_grid_timetable_html(courses, *, title: str) -> str:
    period_time_map = {
        1: "08:10~09:00",
        2: "09:10~10:00",
        3: "10:10~11:00",
        4: "11:10~12:00",
        5: "12:40~13:30",
        6: "13:40~14:30",
        7: "14:40~15:30",
        8: "15:40~16:30",
        9: "16:40~17:30",
        10: "17:40~18:30",
        11: "18:35~19:25",
        12: "19:30~20:20",
        13: "20:25~21:15",
        14: "21:20~22:10",
    }
    day_labels = [("1", "一"), ("2", "二"), ("3", "三"), ("4", "四"), ("5", "五"), ("6", "六"), ("7", "日")]
    period_range = list(range(1, 15))

    timetable = {}
    for c in courses:
        d = safe_str(getattr(c, "day", ""))
        per_raw = safe_str(getattr(c, "period", ""))
        if not d or not per_raw:
            continue
        for p in parse_periods(per_raw):
            timetable.setdefault(d, {}).setdefault(p, []).append(c)

    table_html = '<div class="timetable-wrapper">'
    table_html += f'<div class="timetable-title">{esc(title)}</div>'
    table_html += '<table class="timetable">'
    table_html += "<tr><th>節次</th>"
    for _val, label in day_labels:
        table_html += f"<th>星期{esc(label)}</th>"
    table_html += "</tr>"

    for p in period_range:
        t = period_time_map.get(p, "")
        th_html = f'{p}<div style="font-size:11px;color:#6b7280;margin-top:4px;">{esc(t)}</div>' if t else f"{p}"
        table_html += f"<tr><th>{th_html}</th>"

        for day_val, day_label in day_labels:
            cell_courses = timetable.get(day_val, {}).get(p, [])
            if not cell_courses:
                table_html += "<td>&nbsp;</td>"
                continue

            parts = []
            for c in cell_courses:
                week = safe_str(getattr(c, "week_info", ""))

                # ✅ 修正：data-time 一定要包含數字星期，給前端 parseTimeSlots 用
                period_text = safe_str(getattr(c, "period", ""))
                time_text = f"星期{day_val} 第{period_text}節（星期{day_label}）"
                if week:
                    time_text += f"（{week}）"

                t_ch, t_cat, t_ext = teacher_meta_for_course(c)
                room_txt = room_display(c)
                room_url = building_url_from_room(room_txt)

                parts.append(
                    (
                        f'<div class="course-cell course-clickable" '
                        f'data-id="{c.id}" '
                        f'data-name="{esc(getattr(c, "course_name", ""))}" '
                        f'data-dept="{esc(dept_display(getattr(c, "department_code", "")))}" '
                        f'data-teacher="{esc(getattr(c, "teacher", ""))}" '
                        f'data-teacher-ch="{esc(t_ch)}" '
                        f'data-teacher-category="{esc(t_cat)}" '
                        f'data-teacher-ext="{esc(t_ext)}" '
                        f'data-room="{esc(room_txt)}" '
                        f'data-room-url="{esc(room_url)}" '
                        f'data-time="{esc(time_text)}" '
                        f'data-week="{esc(week)}" '
                        f'data-code="{esc(getattr(c, "course_code", ""))}" '
                        f'data-summary="{esc(getattr(c, "course_summary_ch", ""))}" '
                        f'style="cursor:pointer;">{esc(getattr(c, "course_name", ""))}</div>'
                        f'<div class="course-room">{esc(dept_display(getattr(c, "department_code", "")))}</div>'
                        f'<div class="course-room">{esc(getattr(c, "teacher", ""))} {esc(room_txt)}</div>'
                    )
                )

            table_html += (
                "<td>"
                + "<hr style='border:none;border-top:1px solid #e5e7eb;margin:8px 0;'>".join(parts)
                + "</td>"
            )

        table_html += "</tr>"

    table_html += "</table></div>"
    return table_html


# ==============================
# Profile / Logout
# ==============================
@require_POST
def logout_view(request):
    logout(request)
    return redirect("course_query")


def profile_view(request):
    """處理個人資料管理彈窗送出的『更新密碼』"""
    if request.method == "POST":
        new_password = safe_str(request.POST.get("new_password"))
        confirm_password = safe_str(request.POST.get("confirm_password"))

        if not new_password:
            messages.error(request, "新密碼不能為空白。")
        elif new_password != confirm_password:
            messages.error(request, "新密碼與確認密碼不一致。")
        else:
            user = request.user
            user.set_password(new_password)
            user.save()
            update_session_auth_hash(request, user)
            messages.success(request, "密碼已更新，下一次登入請使用新密碼。")

        return redirect("course_query")

    return redirect("course_query")


# ==============================
# 老師新增/刪除課程
# ==============================
def add_course(request):
    fixed_semester = "1141"
    if not is_teacher_admin(request):
        return redirect("course_query")

    if request.method == "POST":
        form = CourseForm(request.POST)
        if form.is_valid():
            c = form.save(commit=False)
            c.semester = fixed_semester
            c.teacher = "連中岳"
            c.save()
            return redirect(f"{reverse('course_query')}?semester={fixed_semester}&submitted=1")
    else:
        form = CourseForm()

    return render(request, "main/add_course.html", {"form": form})


@require_POST
def delete_course(request, course_id: int):
    if not is_teacher_admin(request):
        return redirect("course_query")

    Course.objects.filter(id=course_id, semester="1141", teacher__icontains="連中岳").delete()
    return redirect(f"{reverse('course_query')}?semester=1141&submitted=1")


# ==============================
# 學生個人課表 AJAX（統一核心邏輯）
# ==============================
def _handle_personal_action(request, action: str, course_id: int, force: bool = False) -> JsonResponse:
    if not is_student_user(request):
        return JsonResponse({"ok": False, "message": "請先以學生身分登入。"}, status=401)

    c = Course.objects.filter(id=course_id).first()
    if not c:
        return JsonResponse({"ok": False, "message": "找不到課程。"}, status=404)

    ensure_fixed_personal_courses(request)
    ids = _get_personal_ids(request)
    id_set = set(ids)
    existing_courses = list(Course.objects.filter(id__in=id_set))
    conflicts = _conflict_slots(existing_courses, c)

    if action == "remove":
        if is_required_course_id(course_id):
            return JsonResponse(
                {"ok": False, "required": True, "message": required_remove_message(course_id)},
                status=409,
            )
        if course_id in id_set:
            ids = [x for x in ids if x != course_id]
            _set_personal_ids(request, ids)
        return JsonResponse({"ok": True, "message": "已從個人課表移除。", "my_course_ids": ids})

    # action == "add"
    if course_id in id_set:
        return JsonResponse({"ok": True, "message": "此課程已在個人課表中。", "my_course_ids": ids})

    if conflicts and not force:
        return JsonResponse(
            {
                "ok": False,
                "conflict": True,
                "conflicts": conflicts,
                "message": f"此課程與你的個人課表衝堂：{_format_conflicts(conflicts)}",
            },
            status=409,
        )

    ids.append(course_id)
    _set_personal_ids(request, ids)
    return JsonResponse(
        {"ok": True, "message": "已新增到個人課表。", "my_course_ids": ids, "warning": bool(conflicts), "conflicts": conflicts}
    )


# ==============================
# 主頁：課程查詢 + 登入 + 顯示
# ==============================
def course_query(request):
    # ✅ 先補齊 demo 帳號（Render 也能用：DEMO_SEED_ACCOUNTS=1 或 DEMO_AUTO_LOGIN=1）
    ensure_default_accounts()

    # ✅ DEMO 一鍵登入（DEMO_AUTO_LOGIN=1 才啟用；並且會補齊 Teacher/Student）
    demo_auto_login(request)

    login_error = ""

    # 0) 先處理「學生 AJAX：新增/移除個人課表」
    if request.method == "POST" and safe_str(request.POST.get("action")) in {"add_my_course", "remove_my_course"}:
        action = safe_str(request.POST.get("action"))
        course_id_raw = safe_str(request.POST.get("course_id"))
        force = safe_str(request.POST.get("force")) == "1"

        try:
            course_id = int(course_id_raw)
        except Exception:
            return JsonResponse({"ok": False, "message": "course_id 格式錯誤。"}, status=400)

        if action == "remove_my_course":
            return _handle_personal_action(request, "remove", course_id, force=force)
        return _handle_personal_action(request, "add", course_id, force=force)

    # 1) 再處理「登入（POST）」：必須 username/password/role 都有效才視為登入
    if request.method == "POST":
        username_in = safe_str(request.POST.get("username"))
        password_in = request.POST.get("password") or ""
        role_in = safe_str(request.POST.get("role"))  # admin / student

        if username_in and password_in and role_in in {"admin", "student"}:
            user = authenticate(request, username=username_in, password=password_in)
            if user is None:
                login_error = "帳號或密碼錯誤"
            else:
                ok = True
                if role_in == "student":
                    # ✅ 這裡不再卡死「一定要先有 Student」，登入成功後會自動補齊
                    ok = True
                elif role_in == "admin":
                    # ✅ 只要是 staff/superuser 或 Teacher 存在就行；登入後也會自動補 Teacher
                    ok = True

                if ok:
                    # ✅ 超關鍵：登入前先補齊 Teacher/Student 綁定，避免「登入了但功能不能用」
                    ensure_role_profile(user, role_in)

                    login(request, user)
                    return redirect("course_query")

    ensure_courses_loaded()

    def norm(v):
        return safe_str(v)

    period_time_map = {
        1: "08:10~09:00",
        2: "09:10~10:00",
        3: "10:10~11:00",
        4: "11:10~12:00",
        5: "12:40~13:30",
        6: "13:40~14:30",
        7: "14:40~15:30",
        8: "15:40~16:30",
        9: "16:40~17:30",
        10: "17:40~18:30",
        11: "18:35~19:25",
        12: "19:30~20:20",
        13: "20:25~21:15",
        14: "21:20~22:10",
    }
    day_labels = [("1", "一"), ("2", "二"), ("3", "三"), ("4", "四"), ("5", "五"), ("6", "六"), ("7", "日")]

    class_type_options = (
        Course.objects.exclude(division__isnull=True)
        .exclude(division__exact="")
        .values_list("division", flat=True)
        .distinct()
        .order_by("division")
    )

    role_name, display_name = get_user_display_name(request.user)
    admin_mode = bool(request.user.is_authenticated and role_name == "老師")

    submitted = request.GET.get("submitted")

    semester = norm(request.GET.get("semester"))
    system = norm(request.GET.get("system"))
    grade = norm(request.GET.get("grade"))
    department = norm(request.GET.get("department"))
    teacher = norm(request.GET.get("teacher"))
    course_name = norm(request.GET.get("course_name"))
    course_code = norm(request.GET.get("course_code"))
    class_type = norm(request.GET.get("class_type"))

    days_selected = request.GET.getlist("day")
    periods_selected = request.GET.getlist("period")

    period_range = list(range(1, 15))
    periods_selected_int = []
    for p in periods_selected:
        try:
            periods_selected_int.append(int(p))
        except Exception:
            continue

    total_count = Course.objects.count()

    # 個人課表（學生）
    personal_timetable_html = ""
    my_course_ids: List[int] = []
    if is_student_user(request):
        ensure_fixed_personal_courses(request)
        my_course_ids = _get_personal_ids(request)
        if my_course_ids:
            m = {c.id: c for c in Course.objects.filter(id__in=my_course_ids)}
            personal_courses = [m[i] for i in my_course_ids if i in m]
            personal_timetable_html = build_grid_timetable_html(personal_courses, title="我的個人課表")
        else:
            personal_timetable_html = '<div class="no-result">找不到 A0 的「系統分析 / 研究概論(資管系)」課程資料。</div>'

    timetable_html = ""

    # ==============================
    # 管理員模式（老師）
    # ==============================
    if admin_mode:
        admin_courses = []
        if submitted and semester:
            qs = Course.objects.filter(teacher__icontains="連中岳", semester=semester)

            def course_sort_key(c):
                return (norm(c.day) or "9", norm(c.period) or "", norm(c.course_name) or "")

            admin_courses = sorted(list(qs), key=course_sort_key)
            for c in admin_courses:
                c.dept_name = dept_display(getattr(c, "department_code", ""))

        context = {
            "semester": semester,
            "system": system,
            "grade": grade,
            "department": department,
            "teacher": teacher,
            "course_name": course_name,
            "course_code": course_code,
            "class_type": class_type,
            "class_type_options": list(class_type_options),
            "days_selected": days_selected,
            "period_range": period_range,
            "periods_selected_int": periods_selected_int,
            "total_count": total_count,
            "timetable_html": "",
            "role_name": role_name,
            "display_name": display_name,
            "login_error": login_error,
            "personal_timetable_html": personal_timetable_html,
            "admin_mode": True,
            "admin_courses": admin_courses,
            "my_course_ids": my_course_ids,
            "current_full_path": request.get_full_path(),
            "building_url_map": json.dumps(BUILDING_URL_MAP),
        }
        return render(request, "main/course_query.html", context)

    # ==============================
    # 學生/一般模式
    # ==============================
    courses = []

    only_semester = bool(semester) and not any(
        [system, grade, department, teacher, course_name, course_code, class_type, days_selected, periods_selected]
    )
    no_condition = not any(
        [semester, system, grade, department, teacher, course_name, course_code, class_type, days_selected, periods_selected]
    )

    if not submitted:
        timetable_html = '<div class="no-result">尚未查詢，請先設定條件後按「查詢」。</div>'
    else:
        if no_condition:
            timetable_html = (
                '<div class="no-result" style="color:#b91c1c;">'
                "請至少選擇一個查詢條件（例如學期、科系、老師、星期或節次）再按「查詢」。"
                "</div>"
            )
        else:
            qs = Course.objects.all()

            if semester:
                qs = qs.filter(semester=semester)
            if system:
                qs = apply_system_filter(qs, system)
            if department:
                qs = qs.filter(department_code__exact=department)
            if grade:
                qs = qs.filter(grade=grade)
            if teacher:
                qs = qs.filter(teacher__icontains=teacher)
            if course_name:
                qs = qs.filter(course_name__icontains=course_name)
            if course_code:
                qs = qs.filter(course_code__icontains=course_code)
            if class_type:
                qs = qs.filter(division__icontains=class_type)
            if days_selected:
                qs = qs.filter(day__in=days_selected)

            courses = list(qs)

            if periods_selected:
                try:
                    period_need = {int(p) for p in periods_selected}
                except Exception:
                    period_need = set()

                if period_need:
                    filtered_list = []
                    for c in courses:
                        c_periods = set(parse_periods(norm(getattr(c, "period", ""))))
                        if c_periods & period_need:
                            filtered_list.append(c)
                    courses = filtered_list

            # 建 timetable dict
            timetable = {}
            for c in courses:
                day_str = norm(c.day)
                periods_raw = norm(c.period)
                if not day_str or not periods_raw:
                    continue
                for pp in parse_periods(periods_raw):
                    timetable.setdefault(day_str, {}).setdefault(pp, []).append(c)

            # 只有學期：顯示列表
            if only_semester:
                if courses:

                    def course_sort_key(c):
                        return (norm(c.day) or "9", norm(c.period), norm(c.department_code), norm(c.class_group))

                    courses_sorted = sorted(courses, key=course_sort_key)
                    day_map = {"1": "一", "2": "二", "3": "三", "4": "四", "5": "五", "6": "六", "7": "日"}

                    list_html = '<div class="timetable-wrapper">'
                    list_html += f'<div class="timetable-title">學期 {esc(semester)} 課程列表</div>'
                    list_html += '<table class="timetable">'
                    list_html += (
                        "<tr>"
                        "<th>課別名稱</th>"
                        "<th>系所</th>"
                        "<th>科目名稱</th>"
                        "<th>老師</th>"
                        "<th>教室地點</th>"
                        "<th>時間</th>"
                        "</tr>"
                    )

                    for c in courses_sorted:
                        dept_name = dept_display(c.department_code)
                        day_num = norm(c.day) or "-"
                        day_ch = day_map.get(day_num, day_num)
                        period_str = norm(c.period) or "-"
                        week_info = norm(c.week_info)

                        # ✅ 顯示用（中文）
                        time_text_display = f"星期{day_ch} 第{period_str}節"
                        if week_info:
                            time_text_display += f"（{week_info}）"

                        # ✅ data-time 用（一定包含數字星期）
                        time_text_data = f"星期{day_num} 第{period_str}節（星期{day_ch}）"
                        if week_info:
                            time_text_data += f"（{week_info}）"

                        t_ch, t_cat, t_ext = teacher_meta_for_course(c)
                        room_txt = room_display(c)
                        room_url = building_url_from_room(room_txt)

                        name_html = (
                            f'<span class="course-clickable" '
                            f'data-id="{c.id}" '
                            f'data-name="{esc(c.course_name)}" '
                            f'data-dept="{esc(dept_name)}" '
                            f'data-teacher="{esc(c.teacher)}" '
                            f'data-teacher-ch="{esc(t_ch)}" '
                            f'data-teacher-category="{esc(t_cat)}" '
                            f'data-teacher-ext="{esc(t_ext)}" '
                            f'data-room="{esc(room_txt)}" '
                            f'data-room-url="{esc(room_url)}" '
                            f'data-time="{esc(time_text_data)}" '
                            f'data-week="{esc(c.week_info)}" '
                            f'data-code="{esc(c.course_code)}" '
                            f'data-summary="{esc(c.course_summary_ch)}" '
                            f'style="cursor:pointer;">{esc(c.course_name)}</span>'
                        )

                        list_html += (
                            "<tr>"
                            f"<td>{esc(c.division) or '-'}</td>"
                            f"<td>{esc(dept_name)}</td>"
                            f"<td>{name_html}</td>"
                            f"<td>{esc(c.teacher) or '-'}</td>"
                            f"<td>{esc(room_txt)}</td>"
                            f"<td>{esc(time_text_display)}</td>"
                            "</tr>"
                        )

                    list_html += "</table></div>"
                    timetable_html = list_html
                else:
                    timetable_html = '<div class="no-result">此學期目前查無任何課程資料。</div>'

            # 多條件：顯示課表 grid
            else:
                if courses:
                    table_html = '<div class="timetable-wrapper">'
                    table_html += '<div class="timetable-title">課表</div>'
                    table_html += '<table class="timetable">'
                    table_html += "<tr><th>節次</th>"
                    for _val, label in day_labels:
                        table_html += f"<th>星期{esc(label)}</th>"
                    table_html += "</tr>"

                    for p in period_range:
                        t = period_time_map.get(p, "")
                        th_html = f'{p}<div style="font-size:11px;color:#6b7280;margin-top:4px;">{esc(t)}</div>' if t else f"{p}"
                        table_html += f"<tr><th>{th_html}</th>"

                        for day_val, day_label in day_labels:
                            if days_selected and day_val not in days_selected:
                                table_html += "<td>&nbsp;</td>"
                                continue

                            courses_in_cell = timetable.get(day_val, {}).get(p, [])
                            if not courses_in_cell:
                                table_html += "<td>&nbsp;</td>"
                                continue

                            # <= 2：直接列
                            if len(courses_in_cell) <= 2:
                                parts = []
                                for c in courses_in_cell:
                                    period_str = norm(c.period) or "-"
                                    week_info = norm(c.week_info)

                                    # ✅ data-time：含數字星期
                                    time_text = f"星期{day_val} 第{period_str}節（星期{day_label}）"
                                    if week_info:
                                        time_text += f"（{week_info}）"

                                    t_ch, t_cat, t_ext = teacher_meta_for_course(c)
                                    room_txt = room_display(c)
                                    room_url = building_url_from_room(room_txt)

                                    parts.append(
                                        (
                                            f'<div class="course-cell course-clickable" '
                                            f'data-id="{c.id}" '
                                            f'data-name="{esc(c.course_name)}" '
                                            f'data-dept="{esc(dept_display(c.department_code))}" '
                                            f'data-teacher="{esc(c.teacher)}" '
                                            f'data-teacher-ch="{esc(t_ch)}" '
                                            f'data-teacher-category="{esc(t_cat)}" '
                                            f'data-teacher-ext="{esc(t_ext)}" '
                                            f'data-room="{esc(room_txt)}" '
                                            f'data-room-url="{esc(room_url)}" '
                                            f'data-time="{esc(time_text)}" '
                                            f'data-week="{esc(c.week_info)}" '
                                            f'data-code="{esc(c.course_code)}" '
                                            f'data-summary="{esc(c.course_summary_ch)}" '
                                            f'style="cursor:pointer;">{esc(c.course_name)}</div>'
                                            f'<div class="course-room">{esc(dept_display(c.department_code))}</div>'
                                            f'<div class="course-room">{esc(c.teacher)} {esc(room_txt)}</div>'
                                        )
                                    )

                                table_html += (
                                    "<td>"
                                    + "<hr style='border:none;border-top:1px solid #e5e7eb;margin:8px 0;'>".join(parts)
                                    + "</td>"
                                )
                                continue

                            # >2：用 select
                            cell_id = f"cell_{day_val}_{p}"
                            first = courses_in_cell[0]

                            select_html = (
                                f'<select class="cell-select" id="{cell_id}_select" '
                                f'onchange="updateTimetableCell(\'{cell_id}\');" '
                                f'title="{esc(first.course_name)}">'
                            )

                            for idx, c in enumerate(courses_in_cell):
                                cname = norm(c.course_name)
                                opt_label = cname if len(cname) <= 18 else cname[:18] + "…"
                                selected = "selected" if idx == 0 else ""

                                t_ch, t_cat, t_ext = teacher_meta_for_course(c)
                                room_txt = room_display(c)
                                room_url = building_url_from_room(room_txt)

                                period_str = norm(getattr(c, "period", "")) or "-"
                                week_info = norm(c.week_info)

                                # ✅ data-time：含數字星期
                                time_text = f"星期{day_val} 第{period_str}節（星期{day_label}）"
                                if week_info:
                                    time_text += f"（{week_info}）"

                                select_html += (
                                    f'<option value="{idx}" {selected} title="{esc(cname)}" '
                                    f'data-id="{c.id}" '
                                    f'data-name="{esc(c.course_name)}" '
                                    f'data-dept="{esc(dept_display(c.department_code))}" '
                                    f'data-teacher="{esc(c.teacher)}" '
                                    f'data-teacher-ch="{esc(t_ch)}" '
                                    f'data-teacher-category="{esc(t_cat)}" '
                                    f'data-teacher-ext="{esc(t_ext)}" '
                                    f'data-room="{esc(room_txt)}" '
                                    f'data-room-url="{esc(room_url)}" '
                                    f'data-time="{esc(time_text)}" '
                                    f'data-week="{esc(c.week_info)}" '
                                    f'data-code="{esc(c.course_code)}" '
                                    f'data-summary="{esc(c.course_summary_ch)}" '
                                    f'>{esc(opt_label)}</option>'
                                )

                            select_html += "</select>"

                            display_html = (
                                f'<div class="cell-display" id="{cell_id}_display">'
                                f'<div class="course-cell">{esc(first.course_name)}</div>'
                                f'<div class="course-room">{esc(dept_display(first.department_code))}</div>'
                                f'<div class="course-room">{esc(first.teacher)} {esc(room_display(first))}</div>'
                                f"</div>"
                            )
                            table_html += f"<td>{select_html}{display_html}</td>"

                        table_html += "</tr>"

                    table_html += "</table></div>"
                    timetable_html = table_html
                else:
                    timetable_html = '<div class="no-result">查無符合條件的課程，請調整查詢條件再試一次。</div>'

    context = {
        "semester": semester,
        "system": system,
        "grade": grade,
        "department": department,
        "teacher": teacher,
        "course_name": course_name,
        "course_code": course_code,
        "class_type": class_type,
        "class_type_options": list(class_type_options),
        "days_selected": days_selected,
        "period_range": period_range,
        "periods_selected_int": periods_selected_int,
        "total_count": total_count,
        "timetable_html": timetable_html,
        "role_name": role_name,
        "display_name": display_name,
        "login_error": login_error,
        "personal_timetable_html": personal_timetable_html,
        "admin_mode": False,
        "my_course_ids": my_course_ids,
        "current_full_path": request.get_full_path(),
        "building_url_map": json.dumps(BUILDING_URL_MAP),
    }
    return render(request, "main/course_query.html", context)


# ==============================
# Excel 匯入工具（安全版）
# ==============================
def _allow_import_via_get() -> bool:
    return bool(getattr(settings, "DEBUG", False))


@login_required
def import_excel(request):
    if not is_teacher_admin(request):
        return HttpResponse("Forbidden", status=403)

    if request.method != "POST" and not _allow_import_via_get():
        return HttpResponse("Method Not Allowed", status=405)

    semester = safe_str(request.GET.get("semester")) or safe_str(request.POST.get("semester")) or "1141"
    file_path = EXCEL_DIR / f"課程查詢_{semester}.xlsx"
    if not file_path.exists():
        return HttpResponse(f"找不到檔案：{file_path}")

    count = _import_xlsx_to_course(file_path)
    return HttpResponse(f"匯入完成（單一檔案 {file_path.name}），共匯入 {count} 筆資料！")


@login_required
def import_all_excels(request):
    if not is_teacher_admin(request):
        return HttpResponse("Forbidden", status=403)

    if request.method != "POST" and not _allow_import_via_get():
        return HttpResponse("Method Not Allowed", status=405)

    excel_files = sorted(EXCEL_DIR.glob("*.xlsx"))
    if not excel_files:
        return HttpResponse(f"在 {EXCEL_DIR} 沒有找到任何 .xlsx 檔案，請確認路徑。")

    total_files = 0
    total_rows = 0
    log_messages = []

    for file_path in excel_files:
        try:
            count = _import_xlsx_to_course(file_path)
            total_files += 1
            total_rows += count
            log_messages.append(f"{file_path.name}：{count} 筆")
        except Exception as e:
            log_messages.append(f"{file_path.name} 匯入失敗：{e}")

    detail = "<br>".join(log_messages)
    return HttpResponse(f"匯入完成，共處理 {total_files} 個檔案，總共 {total_rows} 筆資料。<br><br>{detail}")


# ==============================
# teacher_info（保留）
# ==============================
@require_GET
def teacher_info(request):
    name = safe_str(request.GET.get("name"))
    if not name:
        return JsonResponse({"ok": False, "message": "缺少 name"}, status=400)

    t = Teacher.objects.filter(name_ch=name).first() or Teacher.objects.filter(name_en=name).first()
    if not t:
        return JsonResponse({"ok": False, "message": "找不到教師資料"}, status=404)

    name_ch = safe_str(getattr(t, "name_ch", "")) or name
    category = (
        safe_str(getattr(t, "category", ""))
        or safe_str(getattr(t, "type", ""))
        or safe_str(getattr(t, "title", ""))
        or safe_str(getattr(t, "role", ""))
    )
    ext = (
        safe_str(getattr(t, "extension", ""))
        or safe_str(getattr(t, "ext", ""))
        or safe_str(getattr(t, "phone_ext", ""))
        or safe_str(getattr(t, "office_ext", ""))
    )

    return JsonResponse({"ok": True, "name_ch": name_ch or "-", "category": category or "-", "ext": ext or "-"})


# ==============================
# 回填 classroom（安全版）
# ==============================
@login_required
def backfill_classroom_from_excel(request):
    if not is_teacher_admin(request):
        return HttpResponse("Forbidden", status=403)

    if request.method != "POST" and not _allow_import_via_get():
        return HttpResponse("Method Not Allowed", status=405)

    excel_files = sorted(EXCEL_DIR.glob("*.xlsx"))
    if not excel_files:
        return HttpResponse(f"在 {EXCEL_DIR} 沒有找到任何 .xlsx 檔案。")

    idx = {}
    loaded_rows = 0

    for file_path in excel_files:
        try:
            headers, rows = _iter_xlsx_dict_rows(file_path)
            if not headers:
                continue
        except Exception:
            continue

        for row in rows:
            sem = safe_get(row, "學期")
            code = safe_get(row, "科目代碼(新碼全碼)")
            name = safe_get(row, "科目中文名稱")
            room = room_from_row(row)
            if sem and (code or name) and room:
                idx[(sem, code, name)] = room
                loaded_rows += 1

    qs = Course.objects.filter(Q(classroom__isnull=True) | Q(classroom__exact=""))
    updated = 0

    for c in qs:
        sem = safe_str(getattr(c, "semester", ""))
        code = safe_str(getattr(c, "course_code", ""))
        name = safe_str(getattr(c, "course_name", ""))
        room = idx.get((sem, code, name))
        if room:
            c.classroom = room
            c.save(update_fields=["classroom"])
            updated += 1

    return HttpResponse(f"回填完成：Excel索引 {loaded_rows} 筆；更新 classroom {updated} 筆。")


# ==============================
# 舊網址相容：personal/ & personal/remove
# ==============================
@require_POST
def add_personal_course(request, course_id: int):
    try:
        cid = int(course_id)
    except Exception:
        return JsonResponse({"ok": False, "message": "course_id 格式錯誤。"}, status=400)

    force = safe_str(request.POST.get("force")) == "1"
    return _handle_personal_action(request, "add", cid, force=force)


@require_POST
def remove_personal_course(request, course_id: int):
    try:
        cid = int(course_id)
    except Exception:
        return JsonResponse({"ok": False, "message": "course_id 格式錯誤。"}, status=400)

    return _handle_personal_action(request, "remove", cid, force=False)


# ==============================
# debug_db（建議只在 DEBUG 或 staff 使用）
# ==============================
@require_GET
def debug_db(request):
    if not getattr(settings, "DEBUG", False) and not (request.user.is_authenticated and request.user.is_staff):
        return JsonResponse({"ok": False, "message": "Forbidden"}, status=403)

    return JsonResponse(
        {
            "course_total": Course.objects.count(),
            "semester_distinct": list(
                Course.objects.values_list("semester", flat=True).distinct().order_by("semester")[:50]
            ),
            "sample_3": list(Course.objects.values("id", "semester", "course_name", "teacher", "day", "period")[:3]),
            "excel_dir": str(EXCEL_DIR),
            "excel_files": [p.name for p in sorted(EXCEL_DIR.glob("*.xlsx"))],
            "AUTO_IMPORT": os.environ.get("AUTO_IMPORT", "0"),
            "DEMO_AUTO_LOGIN": os.environ.get("DEMO_AUTO_LOGIN", "0"),
            "DEMO_SEED_ACCOUNTS": os.environ.get("DEMO_SEED_ACCOUNTS", "0"),
        }
    )
@require_GET
def demo_login_view(request):
    """
    ✅ 給 urls.py 用的 demo login endpoint
    - /demo-login/              -> 登入學生 ben
    - /demo-login/?as=teacher   -> 登入老師 dora
    - /demo-login/?logout=1     -> 登出
    """
    ensure_default_accounts()

    if safe_str(request.GET.get("logout")) == "1":
        logout(request)
        return redirect("course_query")

    if request.user.is_authenticated:
        return redirect("course_query")

    as_role = safe_str(request.GET.get("as"))  # "teacher" or "student"
    if as_role == "teacher":
        username = DEFAULT_TEACHER_USERNAME
        role_for_profile = "admin"
    else:
        username = DEFAULT_STUDENT_USERNAME
        role_for_profile = "student"

    User = get_user_model()
    user = User.objects.filter(username=username).first()
    if not user:
        return HttpResponse(
            f"找不到 DEMO 帳號：{username}。請確認已開啟 DEMO_SEED_ACCOUNTS=1 或 DEBUG=True。",
            status=404,
        )

    # ✅ 補齊 Teacher/Student 綁定，避免登入後功能不能用
    ensure_role_profile(user, role_for_profile)

    # ✅ 指定 backend 以支援免密碼登入
    login(request, user, backend="django.contrib.auth.backends.ModelBackend")
    return redirect("course_query")
@require_GET
def demo_logout_view(request):
    """DEMO 登出（給 urls.py 用）"""
    logout(request)
    return redirect("course_query")
